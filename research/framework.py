"""Win-rate engine.

Three functions form the entire public API:

    data              = load_data()
    p_below, p_above  = compute_matrix(data['mvrv'], thresholds, horizons, data)
    write_xlsx(p_below, p_above, OUTPUT / 'research/mvrv.xlsx')

Target throughout: P(price_usd at t+h > price_usd today)
— the direct "am I profitable?" question, no residual detour.

For event-based signals (e.g. WR crossovers) where there is no continuous
feature to threshold, use compute_signal_matrix instead:

    p_bull, p_bear = compute_signal_matrix(
        bull_fn = lambda t: os_reversal(t),   # fn(param) -> bool Series
        bear_fn = lambda t: ob_reversal(t),
        sweep   = range(2, 51, 2),
        horizons = horizons,
        data     = data,
    )
    write_xlsx(p_bull, p_bear, path, sheet_names=('bullish', 'bearish'))
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from typing import Callable
import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import DWD_CSV

MIN_N = 20


# ── data ──────────────────────────────────────────────────────────────────────

def load_data() -> pd.DataFrame:
    """Return the feature warehouse with leak-free expanding-window residual.

    The DWD residual is fitted on all data (look-ahead bias).
    This function replaces it with an expanding-window version where each
    residual value uses only data available on that date.
    """
    data = pd.read_csv(DWD_CSV, index_col='Date', parse_dates=True)
    from research.residual.features import load as _load_residual
    data = data.copy()
    data['log_price_residual'] = _load_residual(data)
    return data


# ── computation ───────────────────────────────────────────────────────────────

def _price_up(data: pd.DataFrame, h: int) -> pd.Series:
    """Boolean: is price h days from now higher than today?"""
    p = data['price_usd'] if 'price_usd' in data.columns else np.exp(data['log_price_usd'])
    return p.shift(-h) > p


def _row(mask: pd.Series, future_ups: dict, horizons: list[int]) -> list:
    """One row of P(price_up) values for a given boolean mask."""
    row, n_val = [], 0
    for h in horizons:
        fu    = future_ups[h]
        valid = mask & fu.notna()
        n     = int(valid.sum())
        p     = float(fu[valid].mean() * 100) if n >= MIN_N else np.nan
        row.append(p)
        if h == horizons[0]:
            n_val = n
    return [n_val] + row


def compute_matrix(
    feature:    pd.Series,
    thresholds: np.ndarray | list,
    horizons:   list[int],
    data:       pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Threshold sweep for a continuous feature.

    Returns (p_below, p_above):
      p_below  rows = thresholds  cols = ['n', '+7d', '+14d', ...]
               value = P(price_up) when feature < threshold
      p_above  same but feature > threshold
    """
    col_labels  = [f'+{h}d' for h in horizons]
    future_ups  = {h: _price_up(data, h) for h in horizons}

    rows_below, rows_above, idx_below, idx_above = [], [], [], []
    for t in thresholds:
        rows_below.append(_row(feature < t, future_ups, horizons))
        rows_above.append(_row(feature > t, future_ups, horizons))
        idx_below.append(f'x < {t:.4g}')
        idx_above.append(f'x > {t:.4g}')

    cols    = ['sample size (n)'] + col_labels
    p_below = pd.DataFrame(rows_below, index=idx_below, columns=cols)
    p_above = pd.DataFrame(rows_above, index=idx_above, columns=cols)
    prob_cols = col_labels
    p_above[prob_cols] = 100 - p_above[prob_cols]   # flip to P(↓)
    p_below.index.name = 'condition'
    p_above.index.name = 'condition'
    return p_below, p_above


def compute_signal_matrix(
    bull_fn:  Callable[[any], pd.Series],
    bear_fn:  Callable[[any], pd.Series],
    sweep:    list,
    horizons: list[int],
    data:     pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parameter sweep for event-based signals.

    bull_fn(param) and bear_fn(param) each return a boolean Series (signal fires).
    sweep is the list of parameter values to iterate over (e.g. range(2, 51, 2)).

    Returns (p_bull, p_bear) with the same shape as compute_matrix output.
    """
    col_labels = [f'+{h}d' for h in horizons]
    future_ups = {h: _price_up(data, h) for h in horizons}

    rows_bull, rows_bear, idx = [], [], []
    for param in sweep:
        rows_bull.append(_row(bull_fn(param).reindex(data.index).fillna(False), future_ups, horizons))
        rows_bear.append(_row(bear_fn(param).reindex(data.index).fillna(False), future_ups, horizons))
        idx.append(str(param))

    cols   = ['sample size (n)'] + col_labels
    p_bull = pd.DataFrame(rows_bull, index=idx, columns=cols)
    p_bear = pd.DataFrame(rows_bear, index=idx, columns=cols)
    p_bull.index.name = p_bear.index.name = 'param'
    return p_bull, p_bear


def compute_condition_matrix(
    conditions: dict[str, pd.Series],
    horizons: list[int],
    data: pd.DataFrame,
) -> pd.DataFrame:
    """Compute P(price_up) for an arbitrary set of named boolean conditions."""
    col_labels = [f'+{h}d' for h in horizons]
    future_ups = {h: _price_up(data, h) for h in horizons}

    rows, idx = [], []
    for name, mask in conditions.items():
        rows.append(_row(mask.reindex(data.index).fillna(False), future_ups, horizons))
        idx.append(name)

    cols = ['sample size (n)'] + col_labels
    out = pd.DataFrame(rows, index=idx, columns=cols)
    out.index.name = 'condition'
    return out


# ── output ────────────────────────────────────────────────────────────────────

_DESCRIPTIONS = {
    '↑': 'P(price at +t days > price now | x < k)',
    '↓': 'P(price at +t days < price now | x > k)',
}


def _write_description(ws, arrow: str) -> None:
    cell = ws['A1']
    cell.value = _DESCRIPTIONS[arrow]
    cell.font  = Font(bold=True, size=20)


def _format_pct(ws, n_rows: int, n_horizons: int, data_row: int = 2) -> None:
    for col in range(3, 3 + n_horizons):
        for row in range(data_row, data_row + n_rows):
            ws.cell(row=row, column=col).number_format = '0.0"%"'


def _color_scale(ws, n_rows: int, n_horizons: int, data_row: int = 2, reverse: bool = False) -> None:
    start = get_column_letter(3)
    end   = get_column_letter(2 + n_horizons)
    lo, hi = ('C00000', '00B050') if not reverse else ('00B050', 'C00000')
    ws.conditional_formatting.add(
        f'{start}{data_row}:{end}{data_row - 1 + n_rows}',
        ColorScaleRule(
            start_type='num', start_value=0,  start_color=lo,
            mid_type='num',   mid_value=50,   mid_color='FFFFFF',
            end_type='num',   end_value=100,  end_color=hi,
        ),
    )


def write_xlsx(
    p_sheet1:    pd.DataFrame,
    p_sheet2:    pd.DataFrame,
    path:        Path,
    sheet_names: tuple[str, str] = ('above', 'below'),
    descriptions: tuple[str, str] | None = None,
) -> None:
    """Write 2-sheet xlsx with description row, then red→white→green colour scale."""
    n_rows     = len(p_sheet1)
    n_horizons = len(p_sheet1.columns) - 1
    descriptions = descriptions or (_DESCRIPTIONS['↑'], _DESCRIPTIONS['↓'])

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # startrow=1 → description in row 1, headers in row 2, data from row 3
        p_sheet1.to_excel(writer, sheet_name=sheet_names[0], startrow=1)
        p_sheet2.to_excel(writer, sheet_name=sheet_names[1], startrow=1)
        for name, description, reverse in zip(sheet_names, descriptions, [False, True]):
            ws = writer.book[name]
            cell = ws['A1']
            cell.value = description
            cell.font = Font(bold=True, size=20)
            _format_pct(ws, n_rows, n_horizons, data_row=3)
            _color_scale(ws, n_rows, n_horizons, data_row=3, reverse=reverse)

    print(f'Saved → {path}')
