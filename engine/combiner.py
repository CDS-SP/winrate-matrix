"""
Naive-Bayes log-odds combination of per-feature conditional probability functions.

For features X_1, ..., X_k assumed conditionally independent given the outcome:
    log-odds(up | x_1, ..., x_k)
        = log-odds(base) + sum_i [ log-odds(P(up | X_i~=x_i)) - log-odds(base) ]

Each term is evaluated from the local finite-difference function embedded in the node xlsx.
"""

import numpy as np
import pandas as pd
import openpyxl

_PDF_SHEET    = 'P(up | X ≈ x) - P(up)'
_PDF_HDR_ROW  = 5   # pandas header at Excel row 5 (startrow=4, empty row 4)
_PDF_DATA_ROW = 6   # data starts at Excel row 6

_SHRINK_FLOOR = 30   # bins with n < 30 contribute zero (CLT floor)
_SHRINK_N0    = 50   # excess observations needed to reach half-weight above the floor


def shrink_weight(n) -> float:
    """
    Two-part shrinkage weight for a bin with n observations.

      n < FLOOR          → 0.0  (below CLT floor; ignore entirely)
      n ≥ FLOOR          → excess / (excess + N0)   where excess = n − FLOOR

    Shape: 0 at n=FLOOR, 0.5 at n=FLOOR+N0, asymptotes to 1 as n→∞.
    """
    n = int(n) if pd.notna(n) else 0
    if n < _SHRINK_FLOOR:
        return 0.0
    excess = n - _SHRINK_FLOOR
    return excess / (excess + _SHRINK_N0)


def _logodds(p_pct: float) -> float:
    p = max(0.001, min(0.999, p_pct / 100.0))
    return float(np.log(p / (1.0 - p)))


def _sigmoid_pct(lo: float) -> float:
    return float(1.0 / (1.0 + np.exp(-lo)) * 100.0)


def load_fn_table(path) -> pd.DataFrame:
    """
    Read the PDF sheet from a node xlsx.
    Returns DataFrame with float midpoint index (x), columns ['n', '+1d', ..., '+14d'].
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[_PDF_SHEET]

    hdr    = next(ws.iter_rows(min_row=_PDF_HDR_ROW, max_row=_PDF_HDR_ROW, values_only=True))
    h_cols = [h for h in hdr if isinstance(h, str) and h.startswith('+')]

    mids, rows = [], []
    for row in ws.iter_rows(min_row=_PDF_DATA_ROW, values_only=True):
        cond = row[0]
        if cond is None:
            break
        try:
            mid = float(str(cond).split('≈')[-1].strip())
        except ValueError:
            continue
        n_val = row[1]
        devs  = [row[i + 2] for i in range(len(h_cols))]
        mids.append(mid)
        rows.append([n_val] + devs)

    wb.close()
    df = pd.DataFrame(rows, index=pd.Index(mids, name='x'), columns=['n'] + h_cols)
    return df.apply(pd.to_numeric, errors='coerce')


def eval_at(fn_table: pd.DataFrame, x: float) -> pd.Series:
    """Linearly interpolate fn_table at feature value x; clamp at range edges."""
    xs = fn_table.index.to_numpy(dtype=float)
    if len(xs) == 0:
        return pd.Series(np.nan, index=fn_table.columns)
    if x <= xs[0]:
        return fn_table.iloc[0]
    if x >= xs[-1]:
        return fn_table.iloc[-1]
    i = int(np.searchsorted(xs, x)) - 1
    t = (x - xs[i]) / (xs[i + 1] - xs[i]) if xs[i + 1] != xs[i] else 0.5
    return fn_table.iloc[i] * (1.0 - t) + fn_table.iloc[i + 1] * t


def peak_signal(fn_table: pd.DataFrame, horizon: str = '+7d', min_n: int = 30) -> float:
    """Max absolute deviation at horizon, restricted to slices with n >= min_n."""
    if horizon not in fn_table.columns:
        return 0.0
    mask = (fn_table['n'] >= min_n) if 'n' in fn_table.columns else pd.Series(True, index=fn_table.index)
    vals = fn_table.loc[mask, horizon].dropna()
    return float(vals.abs().max()) if len(vals) > 0 else 0.0


def combine(
    contributions: list,            # [(node_id: str, devs: pd.Series), ...]
    base_rate:     pd.Series,       # index = horizon labels, values in %
    horizons:      list,
    weights:       dict | None = None,  # {node_id: float} shrinkage weights; default = 1.0 for all
) -> pd.DataFrame:
    """
    Naive-Bayes log-odds combination across signal families.

    Each entry in contributions is (node_id, devs) where devs is a Series
    indexed by horizon label (e.g. '+7d') with values in percentage-point
    deviation from base rate — i.e. P(up | feature ≈ x) − base_rate, as
    stored in the PDF sheet of the node xlsx.

    Formula per horizon h:
        lo = logodds(base_rate_h)
           + Σ_i  w_i · [logodds(base_rate_h + dev_i_h) − logodds(base_rate_h)]
        combined_h = sigmoid(lo) × 100

    Returns DataFrame indexed by horizon labels, columns = ['base_rate'] + node_ids + ['combined', 'edge'].
    """
    node_ids = [nid for nid, _ in contributions]

    records = {}
    for h in horizons:
        br = float(base_rate.loc[h]) if h in base_rate.index else 50.0
        lo = _logodds(br)
        row = {'base_rate': br}
        for nid, devs in contributions:
            dev = float(devs.loc[h]) if h in devs.index else np.nan
            row[nid] = dev
            if pd.notna(dev):
                w = weights[nid] if weights and nid in weights else 1.0
                lo += w * (_logodds(br + dev) - _logodds(br))
        row['combined'] = _sigmoid_pct(lo)
        row['edge']     = row['combined'] - br
        records[h] = row

    col_order = ['base_rate'] + node_ids + ['combined', 'edge']
    return pd.DataFrame(records).T[col_order]
