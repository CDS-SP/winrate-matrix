import numpy as np
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_DEV_FMT  = '+0.0"pp";-0.0"pp";"0pp"'
_SCALE_LO = -20
_SCALE_HI = +20

SHEET_PDF   = 'P(up | X ≈ x) - P(up)'
SHEET_ABOVE = 'P(up | X > x) - P(up)'
SHEET_BELOW = 'P(up | X < x) - P(up)'

_HDR_ROWS  = 3              # rows 1-3: title, formula, description
_START_ROW = _HDR_ROWS + 1  # startrow for to_excel: empty row 4, pandas header row 5
_DATA_ROW  = _HDR_ROWS + 3  # first data row = 6

_MERGE_END = 'P'   # merge A:P across header rows (condition + n + 14 horizons)

_FILL_ROW1 = PatternFill(start_color='666666', end_color='666666', fill_type='solid')
_FILL_ROW2 = PatternFill(start_color='B2B2B2', end_color='B2B2B2', fill_type='solid')
_FILL_ROW3 = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
_CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)

_ABBREVS = {'rsi', 'ma', 'atr', 'dxy', 'macd', 'bb', 'mvrv', 'wr', 'roc', 'vol'}


def _feature_label(node_id: str) -> str:
    words = []
    for p in node_id.split('_'):
        if p.isdigit():
            if words:
                words[-1] = words[-1] + '-' + p
        elif p.lower() in _ABBREVS:
            words.append(p.upper())
        else:
            words.append(p.title())
    return ' '.join(words)


def _subtract_base(df: pd.DataFrame, base_rate: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for col in df.columns:
        if col == 'n':
            continue
        result[col] = df[col] - base_rate.loc[col, 'win_rate']
    return result


def _local_function(p_above: pd.DataFrame, base_rate: pd.DataFrame) -> pd.DataFrame:
    horizon_cols = [c for c in p_above.columns if c != 'n']
    rows, idx = [], []

    for i in range(len(p_above) - 1):
        n_hi    = int(str(p_above.iloc[i]['n']).split('~')[-1])
        n_lo    = int(str(p_above.iloc[i + 1]['n']).split('~')[-1])
        n_slice = n_hi - n_lo
        if n_slice < 1:
            continue

        t_hi = float(p_above.index[i].split()[-1])
        t_lo = float(p_above.index[i + 1].split()[-1])
        mid  = (t_hi + t_lo) / 2

        row = {'n': n_slice}
        for col in horizon_cols:
            v_hi = p_above.iloc[i][col]
            v_lo = p_above.iloc[i + 1][col]
            if pd.notna(v_hi) and pd.notna(v_lo):
                wins_slice = n_hi * v_hi / 100.0 - n_lo * v_lo / 100.0
                row[col]   = wins_slice / n_slice * 100.0 - base_rate.loc[col, 'win_rate']
            else:
                row[col]   = np.nan

        rows.append(row)
        idx.append(f'X ≈ {mid:.4g}')

    result = pd.DataFrame(rows, index=pd.Index(idx, name='condition'))
    return result[['n'] + horizon_cols]


def _color_scale(ws, n_rows: int, n_horizons: int, data_row: int) -> None:
    rng = f'{get_column_letter(3)}{data_row}:{get_column_letter(2 + n_horizons)}{data_row - 1 + n_rows}'
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type='num', start_value=_SCALE_LO, start_color='C00000',
        mid_type='num',   mid_value=0,           mid_color='FFFFFF',
        end_type='num',   end_value=_SCALE_HI,   end_color='00B050',
    ))


def _format_cells(ws, n_rows: int, n_horizons: int, data_row: int) -> None:
    for col in range(3, 3 + n_horizons):
        for row in range(data_row, data_row + n_rows):
            ws.cell(row=row, column=col).number_format = _DEV_FMT


def _write_headers(ws, row1: str, row2: str, row3: str) -> None:
    specs = [
        (1, row1, Font(bold=True,   size=14, color='FFFFFF'), _FILL_ROW1),
        (2, row2, Font(bold=True,   size=13, color='000000'), _FILL_ROW2),
        (3, row3, Font(           size=12, color='000000'), _FILL_ROW3),
    ]
    for r, text, font, fill in specs:
        ws.merge_cells(f'A{r}:{_MERGE_END}{r}')
        cell           = ws[f'A{r}']
        cell.value     = text
        cell.font      = font
        cell.fill      = fill
        cell.alignment = _CENTER
        ws.row_dimensions[r].height = 22


def write_xlsx(
    p_below:   pd.DataFrame,
    p_above:   pd.DataFrame,
    base_rate: pd.DataFrame,
    path:      Path,
    node_id:   str,
) -> None:
    label     = _feature_label(node_id)
    dev_above = _subtract_base(p_above, base_rate)
    dev_below = _subtract_base(p_below, base_rate)
    fn_df     = _local_function(p_above, base_rate)

    n_cdf        = len(dev_above)
    n_pdf        = len(fn_df)
    horizon_cols = [c for c in dev_above.columns if c != 'n']
    n_horizons   = len(horizon_cols)

    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        fn_df.to_excel(    writer, sheet_name=SHEET_PDF,   startrow=_START_ROW)
        dev_above.to_excel(writer, sheet_name=SHEET_ABOVE, startrow=_START_ROW)
        dev_below.to_excel(writer, sheet_name=SHEET_BELOW, startrow=_START_ROW)

        wb = writer.book

        _write_headers(wb[SHEET_PDF],
            f'Conditional Winrate on {label} — Probability Density Function (PDF)',
            f'P(price_h0+h > price_h0 | {label} ≈ x) - P(price_h0+h > price_h0)',
            f'Measures the winrate at horizon h days, given that {label} is approximately at value x, minus the unconditional baseline winrate.',
        )
        _format_cells(wb[SHEET_PDF], n_pdf, n_horizons, data_row=_DATA_ROW)
        _color_scale( wb[SHEET_PDF], n_pdf, n_horizons, data_row=_DATA_ROW)

        _write_headers(wb[SHEET_ABOVE],
            f'Conditional Winrate on {label} — Cumulative Distribution Function (CDF)',
            f'P(price_h0+h > price_h0 | {label} > x) - P(price_h0+h > price_h0)',
            f'Measures the winrate at horizon h days, given that {label} is above threshold x, minus the unconditional baseline winrate.',
        )
        _format_cells(wb[SHEET_ABOVE], n_cdf, n_horizons, data_row=_DATA_ROW)
        _color_scale( wb[SHEET_ABOVE], n_cdf, n_horizons, data_row=_DATA_ROW)

        _write_headers(wb[SHEET_BELOW],
            f'Conditional Winrate on {label} — Cumulative Distribution Function (CDF)',
            f'P(price_h0+h > price_h0 | {label} < x) - P(price_h0+h > price_h0)',
            f'Measures the winrate at horizon h days, given that {label} is below threshold x, minus the unconditional baseline winrate.',
        )
        _format_cells(wb[SHEET_BELOW], n_cdf, n_horizons, data_row=_DATA_ROW)
        _color_scale( wb[SHEET_BELOW], n_cdf, n_horizons, data_row=_DATA_ROW)

    print(f'Saved -> {path}')
